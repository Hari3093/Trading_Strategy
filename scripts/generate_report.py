"""
Trading Journal - Excel Report Generator
Run: python scripts/generate_report.py
Output: trading_report.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import os
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────────
DATA_FILE = "data/trades.csv"
OUTPUT_FILE = "trading_report.xlsx"

# Colors
C_DARK_BG    = "1E1E2E"
C_HEADER     = "2D5A8E"
C_SUBHEADER  = "3A7CBD"
C_GREEN      = "27AE60"
C_RED        = "E74C3C"
C_GOLD       = "F39C12"
C_LIGHT_GRAY = "F5F7FA"
C_MID_GRAY   = "D5D8DC"
C_WHITE      = "FFFFFF"
C_TEXT_DARK  = "2C3E50"
C_WIN_GREEN  = "D5F5E3"
C_LOSS_RED   = "FADBD8"
C_ACCENT     = "8E44AD"

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style='thin', color=C_MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=C_WHITE):
    return Font(name='Arial', bold=bold, size=size, color=color)

def cell_font(size=10, bold=False, color=C_TEXT_DARK):
    return Font(name='Arial', bold=bold, size=size, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center')

def right_align():
    return Alignment(horizontal='right', vertical='center')

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_header_row(ws, row, start_col, end_col, bg=C_HEADER, text=C_WHITE, height=28):
    ws.row_dimensions[row].height = height
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font(color=text)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, start_col, end_col, bg=C_WHITE, height=18):
    ws.row_dimensions[row].height = height
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = cell_font()
        cell.fill = fill(bg)
        cell.border = thin_border()

def kpi_card(ws, row, col, label, value_formula_or_val, fmt="general", label_color=C_HEADER):
    # Label cell
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(name='Arial', bold=True, size=9, color=label_color)
    lc.fill = fill(C_LIGHT_GRAY)
    lc.alignment = center()
    lc.border = thin_border()
    ws.row_dimensions[row].height = 18

    # Value cell
    vc = ws.cell(row=row+1, column=col, value=value_formula_or_val)
    vc.font = Font(name='Arial', bold=True, size=13, color=C_TEXT_DARK)
    vc.fill = fill(C_WHITE)
    vc.alignment = center()
    vc.border = thin_border()
    ws.row_dimensions[row+1].height = 26
    if fmt == "pct":
        vc.number_format = '0.0%'
    elif fmt == "inr":
        vc.number_format = '₹#,##0'
    elif fmt == "x":
        vc.number_format = '0.00"x"'
    elif fmt == "num":
        vc.number_format = '#,##0'

# ─── LOAD & PREP DATA ─────────────────────────────────────────────────────────
def load_data():
    df = pd.read_csv(DATA_FILE, parse_dates=['Date'])
    for col in ['Entry_Price','Exit_Price','Stop_Loss','Take_Profit','Position_Size','Risk_Pct','Profit_Loss']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Month'] = df['Date'].dt.to_period('M')
    df['MonthStr'] = df['Date'].dt.strftime('%b %Y')
    df['Win'] = df['Profit_Loss'] > 0
    return df

# ─── SHEET 1: TRADE LOG ───────────────────────────────────────────────────────
def build_trade_log(wb, df):
    ws = wb.create_sheet("📋 Trade Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:M1")
    tc = ws["A1"]
    tc.value = "📊 TRADING JOURNAL — TRADE LOG"
    tc.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    tc.fill = fill(C_DARK_BG)
    tc.alignment = center()
    ws.row_dimensions[1].height = 36

    # Headers
    headers = [
        "Date", "Time", "Asset", "Type", "Entry ₹", "Exit ₹",
        "Stop Loss", "Take Profit", "Qty", "Risk %", "P&L ₹", "Notes", "Strategy"
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font(size=10)
        c.fill = fill(C_HEADER)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[2].height = 26

    # Data rows
    cols = ['Date','Time','Asset','Trade_Type','Entry_Price','Exit_Price',
            'Stop_Loss','Take_Profit','Position_Size','Risk_Pct','Profit_Loss','Notes','Strategy_Tag']

    for i, (_, row) in enumerate(df.iterrows(), 3):
        is_win = row['Profit_Loss'] > 0
        row_bg = C_WIN_GREEN if is_win else (C_LOSS_RED if row['Profit_Loss'] < 0 else C_WHITE)
        style_data_row(ws, i, 1, 13, bg=row_bg)

        for j, col in enumerate(cols, 1):
            c = ws.cell(row=i, column=j)
            if col == 'Date':
                c.value = row[col].strftime('%d-%b-%Y') if pd.notna(row[col]) else ''
                c.alignment = center()
            elif col == 'Trade_Type':
                c.value = row[col]
                c.font = Font(name='Arial', bold=True, size=10,
                              color=C_GREEN if row[col]=='Buy' else C_RED)
                c.alignment = center()
            elif col in ['Entry_Price','Exit_Price','Stop_Loss','Take_Profit']:
                c.value = row[col]
                c.number_format = '₹#,##0.00'
                c.alignment = right_align()
            elif col == 'Profit_Loss':
                c.value = row[col]
                c.number_format = '₹#,##0;(₹#,##0);-'
                c.font = Font(name='Arial', bold=True, size=10,
                              color=C_GREEN if row[col] > 0 else (C_RED if row[col] < 0 else C_TEXT_DARK))
                c.alignment = right_align()
            elif col == 'Risk_Pct':
                c.value = row[col] / 100
                c.number_format = '0.0%'
                c.alignment = center()
            else:
                c.value = row[col]
                c.alignment = center() if col in ['Time','Position_Size'] else left()

    # Column widths
    set_col_widths(ws, {
        'A':14,'B':8,'C':13,'D':8,'E':11,'F':11,
        'G':11,'H':12,'I':7,'J':8,'K':12,'L':28,'M':16
    })

    # Add filter
    ws.auto_filter.ref = f"A2:M{len(df)+2}"

    # Totals row
    tr = len(df) + 3
    ws.merge_cells(f"A{tr}:J{tr}")
    tc2 = ws.cell(row=tr, column=1, value="TOTALS")
    tc2.font = header_font(size=10, color=C_WHITE)
    tc2.fill = fill(C_SUBHEADER)
    tc2.alignment = center()
    tc2.border = thin_border()
    ws.row_dimensions[tr].height = 22

    total_pnl = ws.cell(row=tr, column=11, value=f"=SUM(K3:K{len(df)+2})")
    total_pnl.number_format = '₹#,##0;(₹#,##0);-'
    total_pnl.font = Font(name='Arial', bold=True, size=11,
                          color=C_GREEN if df['Profit_Loss'].sum() >= 0 else C_RED)
    total_pnl.fill = fill(C_SUBHEADER)
    total_pnl.alignment = center()
    total_pnl.border = thin_border()

    for c in [12, 13]:
        cell = ws.cell(row=tr, column=c)
        cell.fill = fill(C_SUBHEADER)
        cell.border = thin_border()

    return ws

# ─── SHEET 2: PERFORMANCE SUMMARY ─────────────────────────────────────────────
def build_summary(wb, df):
    ws = wb.create_sheet("📈 Performance Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = "📈 PERFORMANCE SUMMARY DASHBOARD"
    tc.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    tc.fill = fill(C_DARK_BG)
    tc.alignment = center()
    ws.row_dimensions[1].height = 36

    # ── KPI Section ──
    ws.merge_cells("A3:L3")
    kh = ws["A3"]
    kh.value = "KEY PERFORMANCE INDICATORS"
    kh.font = header_font(size=11, color=C_WHITE)
    kh.fill = fill(C_HEADER)
    kh.alignment = center()
    ws.row_dimensions[3].height = 24

    # Write raw values for formula references (hidden helper col N onwards)
    total = len(df)
    wins  = int(df['Win'].sum())
    losses = total - wins
    total_pnl = df['Profit_Loss'].sum()
    win_rate = wins / total if total else 0
    avg_win  = df[df['Profit_Loss']>0]['Profit_Loss'].mean() if wins else 0
    avg_loss = df[df['Profit_Loss']<0]['Profit_Loss'].mean() if losses else 0
    rr = abs(avg_win / avg_loss) if avg_loss else 0

    # Running equity for drawdown
    df_sorted = df.sort_values('Date')
    equity = df_sorted['Profit_Loss'].cumsum()
    running_max = equity.cummax()
    drawdown = (equity - running_max)
    max_dd = drawdown.min()

    # KPI Cards — row 4-5
    kpis = [
        ("Total Trades",    total,          "num"),
        ("Winning Trades",  wins,           "num"),
        ("Losing Trades",   losses,         "num"),
        ("Win Rate",        win_rate,       "pct"),
        ("Total P&L",       total_pnl,      "inr"),
        ("Avg Win",         avg_win,        "inr"),
        ("Avg Loss",        avg_loss,       "inr"),
        ("Risk:Reward",     rr,             "x"),
        ("Max Drawdown",    max_dd,         "inr"),
        ("Profit Factor",   abs(df[df['Profit_Loss']>0]['Profit_Loss'].sum() /
                                df[df['Profit_Loss']<0]['Profit_Loss'].sum())
                            if losses else 0, "x"),
        ("Best Trade",      df['Profit_Loss'].max(), "inr"),
        ("Worst Trade",     df['Profit_Loss'].min(), "inr"),
    ]

    for i, (label, value, fmt) in enumerate(kpis):
        col = i + 1
        kpi_card(ws, 4, col, label, value, fmt)

    # ── Strategy Breakdown ──
    ws.merge_cells("A8:F8")
    sh = ws["A8"]
    sh.value = "STRATEGY BREAKDOWN"
    sh.font = header_font(size=11, color=C_WHITE)
    sh.fill = fill(C_ACCENT)
    sh.alignment = center()
    ws.row_dimensions[8].height = 24

    strat_headers = ["Strategy", "Trades", "Wins", "Losses", "Win Rate", "Total P&L"]
    for i, h in enumerate(strat_headers, 1):
        c = ws.cell(row=9, column=i, value=h)
        c.font = header_font(size=10, color=C_WHITE)
        c.fill = fill(C_SUBHEADER)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[9].height = 22

    strat_df = df.groupby('Strategy_Tag').agg(
        Trades=('Profit_Loss','count'),
        Wins=('Win','sum'),
        Total_PnL=('Profit_Loss','sum')
    ).reset_index()
    strat_df['Losses'] = strat_df['Trades'] - strat_df['Wins']
    strat_df['WinRate'] = strat_df['Wins'] / strat_df['Trades']
    strat_df = strat_df.sort_values('Total_PnL', ascending=False)

    for i, (_, row) in enumerate(strat_df.iterrows(), 10):
        bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
        style_data_row(ws, i, 1, 6, bg=bg)
        ws.cell(row=i, column=1, value=row['Strategy_Tag']).alignment = left()
        ws.cell(row=i, column=2, value=row['Trades']).alignment = center()
        ws.cell(row=i, column=3, value=int(row['Wins'])).alignment = center()
        ws.cell(row=i, column=4, value=int(row['Losses'])).alignment = center()
        wr_cell = ws.cell(row=i, column=5, value=row['WinRate'])
        wr_cell.number_format = '0.0%'
        wr_cell.alignment = center()
        pnl_cell = ws.cell(row=i, column=6, value=row['Total_PnL'])
        pnl_cell.number_format = '₹#,##0;(₹#,##0);-'
        pnl_cell.font = Font(name='Arial', size=10, bold=True,
                             color=C_GREEN if row['Total_PnL'] >= 0 else C_RED)
        pnl_cell.alignment = right_align()

    # ── Asset Breakdown ──
    asset_start_col = 8
    ws.merge_cells(f"H8:L8")
    ah = ws["H8"]
    ah.value = "ASSET BREAKDOWN"
    ah.font = header_font(size=11, color=C_WHITE)
    ah.fill = fill(C_GOLD)
    ah.alignment = center()

    asset_headers = ["Asset", "Trades", "Win Rate", "Total P&L", "Avg P&L"]
    for i, h in enumerate(asset_headers, asset_start_col):
        c = ws.cell(row=9, column=i, value=h)
        c.font = header_font(size=10, color=C_WHITE)
        c.fill = fill(C_SUBHEADER)
        c.alignment = center()
        c.border = thin_border()

    asset_df = df.groupby('Asset').agg(
        Trades=('Profit_Loss','count'),
        Wins=('Win','sum'),
        Total_PnL=('Profit_Loss','sum'),
        Avg_PnL=('Profit_Loss','mean')
    ).reset_index()
    asset_df['WinRate'] = asset_df['Wins'] / asset_df['Trades']
    asset_df = asset_df.sort_values('Total_PnL', ascending=False)

    for i, (_, row) in enumerate(asset_df.iterrows(), 10):
        bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
        style_data_row(ws, i, asset_start_col, asset_start_col+4, bg=bg)
        ws.cell(row=i, column=asset_start_col,   value=row['Asset']).alignment = center()
        ws.cell(row=i, column=asset_start_col+1, value=row['Trades']).alignment = center()
        wr = ws.cell(row=i, column=asset_start_col+2, value=row['WinRate'])
        wr.number_format = '0.0%'; wr.alignment = center()
        pnl = ws.cell(row=i, column=asset_start_col+3, value=row['Total_PnL'])
        pnl.number_format = '₹#,##0;(₹#,##0);-'
        pnl.font = Font(name='Arial', size=10, bold=True,
                        color=C_GREEN if row['Total_PnL'] >= 0 else C_RED)
        pnl.alignment = right_align()
        avg = ws.cell(row=i, column=asset_start_col+4, value=row['Avg_PnL'])
        avg.number_format = '₹#,##0;(₹#,##0);-'
        avg.alignment = right_align()

    # ── Equity Curve Data (for chart) ──
    eq_row_start = 22
    ws.merge_cells(f"A{eq_row_start}:F{eq_row_start}")
    eqh = ws[f"A{eq_row_start}"]
    eqh.value = "EQUITY CURVE DATA"
    eqh.font = header_font(size=11, color=C_WHITE)
    eqh.fill = fill(C_HEADER)
    eqh.alignment = center()
    ws.row_dimensions[eq_row_start].height = 24

    eq_headers = ["Trade #", "Date", "Asset", "P&L", "Cumulative P&L", "Drawdown"]
    for i, h in enumerate(eq_headers, 1):
        c = ws.cell(row=eq_row_start+1, column=i, value=h)
        c.font = header_font(size=10, color=C_WHITE)
        c.fill = fill(C_SUBHEADER)
        c.alignment = center()
        c.border = thin_border()

    df_sorted = df.sort_values('Date').reset_index(drop=True)
    cum_pnl = 0
    for i, (_, row) in enumerate(df_sorted.iterrows(), eq_row_start+2):
        cum_pnl += row['Profit_Loss']
        bg = C_LIGHT_GRAY if i % 2 == 0 else C_WHITE
        style_data_row(ws, i, 1, 6, bg=bg)
        ws.cell(row=i, column=1, value=i - eq_row_start - 1).alignment = center()
        ws.cell(row=i, column=2, value=row['Date'].strftime('%d-%b')).alignment = center()
        ws.cell(row=i, column=3, value=row['Asset']).alignment = center()
        p = ws.cell(row=i, column=4, value=row['Profit_Loss'])
        p.number_format = '₹#,##0;(₹#,##0);-'
        p.alignment = right_align()
        cp = ws.cell(row=i, column=5, value=cum_pnl)
        cp.number_format = '₹#,##0;(₹#,##0);-'
        cp.alignment = right_align()
        dd_val = drawdown.iloc[i - eq_row_start - 2] if (i - eq_row_start - 2) < len(drawdown) else 0
        dd = ws.cell(row=i, column=6, value=dd_val)
        dd.number_format = '₹#,##0;(₹#,##0);-'
        dd.alignment = right_align()

    # Equity Curve Chart
    chart_data_end = eq_row_start + 1 + len(df)
    line = LineChart()
    line.title = "Equity Curve"
    line.style = 10
    line.y_axis.title = "Cumulative P&L (₹)"
    line.x_axis.title = "Trade #"
    line.height = 12
    line.width = 22

    data = Reference(ws, min_col=5, min_row=eq_row_start+1, max_row=chart_data_end)
    line.add_data(data, titles_from_data=True)
    line.series[0].graphicalProperties.line.solidFill = C_GREEN
    line.series[0].graphicalProperties.line.width = 20000

    ws.add_chart(line, "H22")

    # Win/Loss Pie Chart
    pie = PieChart()
    pie.title = "Win / Loss Ratio"
    pie.height = 10
    pie.width = 14

    pie_data_row = chart_data_end + 2
    ws.cell(row=pie_data_row, column=1, value="Result")
    ws.cell(row=pie_data_row, column=2, value="Count")
    ws.cell(row=pie_data_row+1, column=1, value="Wins")
    ws.cell(row=pie_data_row+1, column=2, value=wins)
    ws.cell(row=pie_data_row+2, column=1, value="Losses")
    ws.cell(row=pie_data_row+2, column=2, value=losses)

    pie_data = Reference(ws, min_col=2, min_row=pie_data_row, max_row=pie_data_row+2)
    pie_labels = Reference(ws, min_col=1, min_row=pie_data_row+1, max_row=pie_data_row+2)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    slice_win = DataPoint(idx=0)
    slice_win.graphicalProperties.solidFill = C_GREEN
    slice_loss = DataPoint(idx=1)
    slice_loss.graphicalProperties.solidFill = C_RED
    pie.series[0].data_points = [slice_win, slice_loss]
    ws.add_chart(pie, "A38")

    # Column widths
    set_col_widths(ws, {
        'A':14,'B':10,'C':10,'D':10,'E':10,'F':11,
        'G':4,'H':14,'I':10,'J':10,'K':12,'L':12
    })

    return ws

# ─── SHEET 3: MONTHLY STATS ────────────────────────────────────────────────────
def build_monthly(wb, df):
    ws = wb.create_sheet("📅 Monthly Stats")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    tc = ws["A1"]
    tc.value = "📅 MONTHLY PERFORMANCE STATISTICS"
    tc.font = Font(name='Arial', bold=True, size=14, color=C_WHITE)
    tc.fill = fill(C_DARK_BG)
    tc.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["Month","Trades","Wins","Losses","Win Rate","Gross Profit","Gross Loss","Net P&L","Avg Trade","Best Trade"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font(size=10, color=C_WHITE)
        c.fill = fill(C_HEADER)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[2].height = 26

    monthly = df.groupby('MonthStr').agg(
        Trades=('Profit_Loss','count'),
        Wins=('Win','sum'),
        Gross_Profit=('Profit_Loss', lambda x: x[x>0].sum()),
        Gross_Loss=('Profit_Loss',   lambda x: x[x<0].sum()),
        Net_PnL=('Profit_Loss','sum'),
        Avg_Trade=('Profit_Loss','mean'),
        Best_Trade=('Profit_Loss','max'),
    ).reset_index()
    monthly['Losses'] = monthly['Trades'] - monthly['Wins']
    monthly['Win_Rate'] = monthly['Wins'] / monthly['Trades']

    # Sort by date
    monthly['_sort'] = pd.to_datetime(monthly['MonthStr'], format='%b %Y')
    monthly = monthly.sort_values('_sort').drop(columns='_sort')

    for i, (_, row) in enumerate(monthly.iterrows(), 3):
        is_pos = row['Net_PnL'] >= 0
        bg = C_WIN_GREEN if is_pos else C_LOSS_RED
        style_data_row(ws, i, 1, 10, bg=bg)

        ws.cell(row=i, column=1, value=row['MonthStr']).alignment = center()
        ws.cell(row=i, column=2, value=row['Trades']).alignment = center()
        ws.cell(row=i, column=3, value=int(row['Wins'])).alignment = center()
        ws.cell(row=i, column=4, value=int(row['Losses'])).alignment = center()

        wr = ws.cell(row=i, column=5, value=row['Win_Rate'])
        wr.number_format = '0.0%'; wr.alignment = center()

        for col, val in [(6, row['Gross_Profit']), (7, row['Gross_Loss']),
                          (8, row['Net_PnL']),      (9, row['Avg_Trade']),
                          (10, row['Best_Trade'])]:
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = '₹#,##0;(₹#,##0);-'
            c.alignment = right_align()
            if col == 8:
                c.font = Font(name='Arial', bold=True, size=10,
                              color=C_GREEN if val >= 0 else C_RED)

    # Totals
    tr = len(monthly) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = header_font(size=10, color=C_WHITE)
    ws.cell(row=tr, column=1).fill = fill(C_HEADER)
    ws.cell(row=tr, column=1).alignment = center()
    ws.cell(row=tr, column=1).border = thin_border()

    for col in range(2, 11):
        c = ws.cell(row=tr, column=col)
        c.fill = fill(C_HEADER)
        c.border = thin_border()
        c.alignment = center()
        if col in [2,3,4]:
            c.value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})"
            c.font = header_font(size=10)
        elif col == 5:
            c.value = f"=C{tr}/B{tr}"
            c.number_format = '0.0%'
            c.font = header_font(size=10)
        elif col in [6,7,8]:
            c.value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr-1})"
            c.number_format = '₹#,##0;(₹#,##0);-'
            c.font = header_font(size=10)
        elif col == 9:
            c.value = f"=H{tr}/B{tr}"
            c.number_format = '₹#,##0;(₹#,##0);-'
            c.font = header_font(size=10)
        elif col == 10:
            c.value = f"=MAX(J3:J{tr-1})"
            c.number_format = '₹#,##0;(₹#,##0);-'
            c.font = header_font(size=10)
    ws.row_dimensions[tr].height = 22

    # Monthly P&L Bar Chart
    chart_start = len(monthly) + 6
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Monthly Net P&L"
    bar.y_axis.title = "P&L (₹)"
    bar.x_axis.title = "Month"
    bar.height = 13
    bar.width = 26

    data = Reference(ws, min_col=8, min_row=2, max_row=len(monthly)+2)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(monthly)+2)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = C_HEADER
    ws.add_chart(bar, f"A{chart_start}")

    # Monthly Win Rate Line Chart
    line = LineChart()
    line.title = "Monthly Win Rate"
    line.y_axis.title = "Win Rate"
    line.x_axis.title = "Month"
    line.height = 13
    line.width = 18

    wr_data = Reference(ws, min_col=5, min_row=2, max_row=len(monthly)+2)
    line.add_data(wr_data, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties.line.solidFill = C_GOLD
    line.series[0].graphicalProperties.line.width = 20000
    ws.add_chart(line, f"K{chart_start}")

    set_col_widths(ws, {
        'A':13,'B':9,'C':8,'D':9,'E':10,
        'F':14,'G':14,'H':12,'I':12,'J':13
    })

    return ws

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("📊 Trading Journal — Generating Excel Report...")
    df = load_data()
    print(f"   ✅ Loaded {len(df)} trades from {DATA_FILE}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_trade_log(wb, df)
    print("   ✅ Sheet 1: Trade Log built")

    build_summary(wb, df)
    print("   ✅ Sheet 2: Performance Summary built")

    build_monthly(wb, df)
    print("   ✅ Sheet 3: Monthly Stats built")

    # Set tab colors
    wb["📋 Trade Log"].sheet_properties.tabColor          = "2D5A8E"
    wb["📈 Performance Summary"].sheet_properties.tabColor = "27AE60"
    wb["📅 Monthly Stats"].sheet_properties.tabColor       = "F39C12"

    wb.save(OUTPUT_FILE)
    print(f"\n🎉 Done! Report saved: {OUTPUT_FILE}")
    print(f"   📁 Open it in Excel or Google Sheets")

if __name__ == "__main__":
    main()
